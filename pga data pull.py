import pandas as pd
import openpyxl, pprint

#combines all of the 5 years of PGA data into a dictionary with keys being name of player
#player age and other unnecessary symbols are taken out

Edit = {}
PGA = {}
wb = openpyxl.load_workbook('PGAData.xlsx')
sheet = wb['Sheet1']
for row in range(2, sheet.max_row + 1 ):
    Edit[sheet['A' + str(row)].value] = [sheet['B' + str(row)].value, sheet['C' + str(row)].value]

for n in Edit:
    event = Edit[n][0]
    sal = Edit[n][1]
    for let in n:
        if let in '1234567890, ':
            editname = n.replace(let,"")
            n = editname
    PGA[editname] = [event,sal]
    
    
for k in PGA:
    for element in PGA[k][1]:
        if element in ',$':
            PGA[k][1] = PGA[k][1].replace(element,"")  
    PGA[k][1] = float(PGA[k][1])
    


Edit2 = {}
PGA2 = {}
wb2 = openpyxl.load_workbook('PGAData16.xlsx')
sheet2 = wb2['Sheet1']
for row in range(2, sheet2.max_row + 1 ):
    Edit2[sheet2['A' + str(row)].value] = [sheet2['B' + str(row)].value, sheet2['C' + str(row)].value]

for n in Edit2:
    event = Edit2[n][0]
    sal = Edit2[n][1]
    for let in n:
        if let in '1234567890, ':
            editname = n.replace(let,"")
            n = editname
    PGA2[editname] = [event,sal]
    
    
for k in PGA2:
    for element in PGA2[k][1]:
        if element in ',$':
            PGA2[k][1] = PGA2[k][1].replace(element,"")  
    PGA2[k][1] = float(PGA2[k][1])


Edit3 = {}
PGA3 = {}
wb3 = openpyxl.load_workbook('PGAData17.xlsx')
sheet3 = wb3['Sheet1']
for row in range(2, sheet3.max_row + 1 ):
    Edit3[sheet3['A' + str(row)].value] = [sheet3['B' + str(row)].value, sheet3['C' + str(row)].value]

for n in Edit3:
    event = Edit3[n][0]
    sal = Edit3[n][1]
    for let in n:
        if let in '1234567890, ':
            editname = n.replace(let,"")
            n = editname
    PGA3[editname] = [event,sal]
    
    
for k in PGA3:
    for element in PGA3[k][1]:
        if element in ',$':
            PGA3[k][1] = PGA3[k][1].replace(element,"")  
    PGA3[k][1] = float(PGA3[k][1])
    
    
Edit4 = {}
PGA4 = {}
wb4 = openpyxl.load_workbook('PGAData18.xlsx')
sheet4 = wb4['Sheet1']
for row in range(2, sheet4.max_row + 1 ):
    Edit4[sheet4['A' + str(row)].value] = [sheet4['B' + str(row)].value, sheet4['C' + str(row)].value]

for n in Edit4:
    event = Edit4[n][0]
    sal = Edit4[n][1]
    for let in n:
        if let in '1234567890, ':
            editname = n.replace(let,"")
            n = editname
    PGA4[editname] = [event,sal]
    
    
for k in PGA4:
    for element in PGA4[k][1]:
        if element in ',$':
            PGA4[k][1] = PGA4[k][1].replace(element,"")  
    PGA4[k][1] = float(PGA4[k][1])
    
Edit5 = {}
PGA5 = {}
wb5 = openpyxl.load_workbook('PGAData19.xlsx')
sheet5 = wb5['Sheet1']
for row in range(2, sheet5.max_row + 1 ):
    Edit5[sheet5['A' + str(row)].value] = [sheet5['B' + str(row)].value, sheet5['C' + str(row)].value]

for n in Edit5:
    event = Edit5[n][0]
    sal = Edit5[n][1]
    for let in n:
        if let in '1234567890, ':
            editname = n.replace(let,"")
            n = editname
    PGA5[editname] = [event,sal]
    
    
for k in PGA5:
    for element in PGA5[k][1]:
        if element in ',$':
            PGA5[k][1] = PGA5[k][1].replace(element,"")  
    PGA5[k][1] = float(PGA5[k][1])
    
    
    
dct = [PGA2,PGA3,PGA4,PGA5]
for n in dct:
    for key in n:
        evt = n[key][0]
        sal = n[key][1]
        if key in PGA:
            PGA[key][0] = PGA[key][0] + evt
            PGA[key][1] = PGA[key][1] + sal
            
        else:
            PGA[key] = [evt,sal]
            

import csv


w = csv.writer(open("CompletedPGA.csv", "w"))
for key, val in PGA.items():
    w.writerow([key,val[0],val[1]])
    